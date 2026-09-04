import io
import json
import uuid
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session

from backend.app.models.merchant import Merchant
from backend.app.models.product import Product
from backend.app.models.sale import Sale, SaleItem


class AnalyticsService:
    def get_period_sales_analytics(
        self,
        db: Session,
        merchant_id: uuid.UUID,
    ) -> Dict[str, Any]:
        """
        Calculates sales and product metrics segmented by Today (Day), This Week (7 Days),
        This Month (30 Days), and All-Time, strictly isolated by merchant_id.
        """
        merchant = db.query(Merchant).filter(Merchant.id == merchant_id).first()
        if not merchant:
            return {}

        now_utc = datetime.now(timezone.utc)
        today_start = datetime(now_utc.year, now_utc.month, now_utc.day, tzinfo=timezone.utc)
        week_start = now_utc - timedelta(days=7)
        month_start = now_utc - timedelta(days=30)

        all_sales = (
            db.query(Sale)
            .filter(Sale.merchant_id == merchant.id)
            .order_by(Sale.created_at.desc())
            .all()
        )
        products = db.query(Product).filter(Product.merchant_id == merchant.id).all()

        def _compute_period_stats(sales_list: List[Sale]) -> Dict[str, Any]:
            count = len(sales_list)
            gmv = sum(s.total_amount for s in sales_list)
            collected = sum(s.received_amount for s in sales_list)
            outstanding = sum(s.outstanding_amount for s in sales_list)
            paid_count = sum(1 for s in sales_list if s.status == "PAID")
            pending_count = sum(1 for s in sales_list if s.status == "PENDING")
            partial_count = sum(1 for s in sales_list if s.status == "PARTIAL")
            collection_rate = (collected / gmv * 100.0) if gmv > 0 else 100.0

            # Count product units and revenue in this period
            product_sales: Dict[str, Dict[str, Any]] = {}
            for s in sales_list:
                for it in s.items:
                    p_name = it.product_name.lower().strip()
                    if p_name not in product_sales:
                        product_sales[p_name] = {"units": 0, "revenue": 0.0}
                    product_sales[p_name]["units"] += it.quantity
                    product_sales[p_name]["revenue"] += it.subtotal

            top_products = [
                {"name": name, "units": data["units"], "revenue": round(data["revenue"], 2)}
                for name, data in sorted(product_sales.items(), key=lambda x: x[1]["revenue"], reverse=True)
            ]

            return {
                "orders_count": count,
                "total_gmv": round(gmv, 2),
                "total_collected": round(collected, 2),
                "total_outstanding": round(outstanding, 2),
                "paid_orders_count": paid_count,
                "pending_orders_count": pending_count,
                "partial_orders_count": partial_count,
                "collection_rate": round(collection_rate, 1),
                "top_products": top_products[:5],
                "product_sales_map": product_sales,
            }

        sales_today = [s for s in all_sales if s.created_at >= today_start]
        sales_week = [s for s in all_sales if s.created_at >= week_start]
        sales_month = [s for s in all_sales if s.created_at >= month_start]

        today_stats = _compute_period_stats(sales_today)
        week_stats = _compute_period_stats(sales_week)
        month_stats = _compute_period_stats(sales_month)
        all_time_stats = _compute_period_stats(all_sales)

        # Catalog performance summary
        catalog_summary = []
        for p in products:
            attrs = p.attributes if isinstance(p.attributes, dict) else {}
            p_name = p.name.lower().strip()
            month_p = month_stats["product_sales_map"].get(p_name, {"units": 0, "revenue": 0.0})
            all_p = all_time_stats["product_sales_map"].get(p_name, {"units": 0, "revenue": 0.0})

            catalog_summary.append({
                "id": str(p.id),
                "name": p.name,
                "category": p.category or "General",
                "price": p.price,
                "unit": p.unit or "piece",
                "stock_quantity": p.stock_quantity,
                "description": p.description or "",
                "attributes": attrs,
                "is_active": p.is_active,
                "units_sold_month": month_p["units"],
                "units_sold_all_time": all_p["units"],
                "revenue_all_time": round(all_p["revenue"], 2),
            })

        return {
            "merchant": {
                "id": str(merchant.id),
                "name": merchant.name,
                "business_type": merchant.business_type,
                "currency": merchant.currency or "INR",
            },
            "generated_at": now_utc.isoformat(),
            "periods": {
                "today": today_stats,
                "week": week_stats,
                "month": month_stats,
                "all_time": all_time_stats,
            },
            "catalog_summary": catalog_summary,
        }

    def generate_excel_report(self, db: Session, merchant_id: uuid.UUID) -> bytes:
        """
        Generates a professionally styled Excel workbook (.xlsx) with 3 sheets:
        1. Executive Sales Analytics (Day / Week / Month Summary)
        2. Product Catalog & Performance
        3. Detailed Sales Transactions Ledger
        """
        analytics = self.get_period_sales_analytics(db, merchant_id)
        merchant_info = analytics.get("merchant", {})

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl is required to generate Excel reports. Please install it.")

        all_sales = (
            db.query(Sale)
            .filter(Sale.merchant_id == merchant_id)
            .order_by(Sale.created_at.desc())
            .all()
        )

        wb = openpyxl.Workbook()

        # Styles
        font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)
        font_sub = Font(name="Calibri", size=10, italic=True, color="555555")

        fill_primary = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")     # Indigo
        fill_secondary = PatternFill(start_color="059669", end_color="059669", fill_type="solid")   # Emerald
        fill_dark = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")        # Slate Dark
        fill_row_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_paid = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fill_pending = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        fill_partial = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        # ─────────────────────────────────────────────────────────────
        # SHEET 1: EXECUTIVE SUMMARY (Day / Week / Month)
        # ─────────────────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Sales Analytics Summary"
        ws1.views.sheetView[0].showGridLines = True

        # Header Title Banner
        ws1.merge_cells("A1:G2")
        title_cell = ws1["A1"]
        title_cell.value = f"📊 {merchant_info.get('name', 'Store')} — Sales & Performance Report"
        title_cell.font = font_title
        title_cell.fill = fill_primary
        title_cell.alignment = align_center

        ws1["A3"] = f"Business Type: {merchant_info.get('business_type', 'Retail')}  |  Currency: {merchant_info.get('currency', 'INR')}  |  Generated on: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
        ws1["A3"].font = font_sub

        # Period Table Headers
        period_headers = [
            "Period",
            "Orders Count",
            f"Gross Sales ({merchant_info.get('currency', 'INR')})",
            f"Collected Amount ({merchant_info.get('currency', 'INR')})",
            f"Outstanding Amount ({merchant_info.get('currency', 'INR')})",
            "Paid Orders",
            "Collection Rate (%)",
        ]

        row_idx = 5
        for col_idx, h in enumerate(period_headers, 1):
            c = ws1.cell(row=row_idx, column=col_idx, value=h)
            c.font = font_header
            c.fill = fill_dark
            c.alignment = align_center
            c.border = thin_border

        periods_data = [
            ("Today (Last 24h)", analytics["periods"]["today"]),
            ("This Week (Last 7 Days)", analytics["periods"]["week"]),
            ("This Month (Last 30 Days)", analytics["periods"]["month"]),
            ("All-Time Cumulative", analytics["periods"]["all_time"]),
        ]

        for p_label, p_data in periods_data:
            row_idx += 1
            is_all_time = "Cumulative" in p_label
            row_font = font_bold if is_all_time else font_regular
            bg = fill_row_zebra if row_idx % 2 == 0 and not is_all_time else PatternFill(fill_type=None)

            row_values = [
                p_label,
                p_data["orders_count"],
                p_data["total_gmv"],
                p_data["total_collected"],
                p_data["total_outstanding"],
                f"{p_data['paid_orders_count']} / {p_data['orders_count']}",
                f"{p_data['collection_rate']}%",
            ]

            for col_idx, val in enumerate(row_values, 1):
                c = ws1.cell(row=row_idx, column=col_idx, value=val)
                c.font = row_font
                c.border = thin_border
                if not is_all_time and bg.fill_type:
                    c.fill = bg
                c.alignment = align_left if col_idx == 1 else align_right

        # Auto-fit columns
        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws1.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 15)

        # ─────────────────────────────────────────────────────────────
        # SHEET 2: PRODUCT CATALOG & VOLUME
        # ─────────────────────────────────────────────────────────────
        ws2 = wb.create_sheet(title="Catalog & Products")
        ws2.views.sheetView[0].showGridLines = True

        ws2.merge_cells("A1:H2")
        t2 = ws2["A1"]
        t2.value = f"📦 {merchant_info.get('name', 'Store')} — Catalog & Inventory Inventory"
        t2.font = font_title
        t2.fill = fill_secondary
        t2.alignment = align_center

        cat_headers = [
            "Item Name",
            "Category",
            f"Price ({merchant_info.get('currency', 'INR')})",
            "Unit",
            "Stock Qty",
            "Units Sold (30 Days)",
            "Units Sold (All-Time)",
            f"Total Revenue ({merchant_info.get('currency', 'INR')})",
        ]

        row_idx = 4
        for col_idx, h in enumerate(cat_headers, 1):
            c = ws2.cell(row=row_idx, column=col_idx, value=h)
            c.font = font_header
            c.fill = fill_dark
            c.alignment = align_center
            c.border = thin_border

        for item in analytics.get("catalog_summary", []):
            row_idx += 1
            bg = fill_row_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
            vals = [
                item["name"].title(),
                item["category"],
                item["price"],
                item["unit"],
                item.get("stock_quantity", 0),
                item["units_sold_month"],
                item["units_sold_all_time"],
                item["revenue_all_time"],
            ]
            for col_idx, val in enumerate(vals, 1):
                c = ws2.cell(row=row_idx, column=col_idx, value=val)
                c.font = font_regular
                c.border = thin_border
                if bg.fill_type:
                    c.fill = bg
                c.alignment = align_left if col_idx in (1, 2, 4) else align_right

        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 15)

        # ─────────────────────────────────────────────────────────────
        # SHEET 3: TRANSACTIONS & SALES LEDGER
        # ─────────────────────────────────────────────────────────────
        ws3 = wb.create_sheet(title="Sales Transactions Ledger")
        ws3.views.sheetView[0].showGridLines = True

        ws3.merge_cells("A1:H2")
        t3 = ws3["A1"]
        t3.value = f"🧾 {merchant_info.get('name', 'Store')} — Sales Transactions Ledger"
        t3.font = font_title
        t3.fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
        t3.alignment = align_center

        txn_headers = [
            "Order ID",
            "Date & Time",
            "Customer Name",
            "Items Purchased",
            f"Total ({merchant_info.get('currency', 'INR')})",
            f"Received ({merchant_info.get('currency', 'INR')})",
            f"Outstanding ({merchant_info.get('currency', 'INR')})",
            "Status",
        ]

        row_idx = 4
        for col_idx, h in enumerate(txn_headers, 1):
            c = ws3.cell(row=row_idx, column=col_idx, value=h)
            c.font = font_header
            c.fill = fill_dark
            c.alignment = align_center
            c.border = thin_border

        for s in all_sales:
            row_idx += 1
            item_desc = ", ".join(f"{it.quantity}x {it.product_name}" for it in s.items) or "N/A"
            date_str = s.created_at.strftime("%d %b %Y, %H:%M") if s.created_at else "N/A"

            vals = [
                s.id,
                date_str,
                s.customer_name or "Walk-in Customer",
                item_desc,
                s.total_amount,
                s.received_amount,
                s.outstanding_amount,
                s.status,
            ]

            for col_idx, val in enumerate(vals, 1):
                c = ws3.cell(row=row_idx, column=col_idx, value=val)
                c.font = font_regular
                c.border = thin_border
                c.alignment = align_left if col_idx in (1, 2, 3, 4) else align_right

                if col_idx == 8:
                    c.alignment = align_center
                    if s.status == "PAID":
                        c.fill = fill_paid
                    elif s.status == "PARTIAL":
                        c.fill = fill_partial
                    else:
                        c.fill = fill_pending

        for col in ws3.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws3.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 15)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()


analytics_service = AnalyticsService()
